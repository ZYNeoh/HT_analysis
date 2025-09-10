"""
This file is for execution
"""
import numpy as np 
import matplotlib.pyplot as plt
import random
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from playsound import playsound

import core_HT
import material
import Validation

def HT_data_export_preparation (selection, value_1) :
    data_x = []
    data_y = []
    data_z = []


    if selection == 3 :
        data_x = [0]
        data_y = [value_1]
    elif selection == 5 :
        data_x = [0]
        data_y = [0]
    elif selection == 100 :
        data_x = 0
    return data_x, data_y, data_z

def HT_data_export (
        export_data_x, export_data_y, export_data_z, selection, x_data_2, x_data,
        fluid_h, As, r_set, Q_dot, Te, init_T_i, T_i, velo_fluid_new, 
        m_dot, fluid_cp
) :
    if selection == 0 :
        r_set_2 = [1/fluid_h/As]
        r_set_2.extend(r_set)
        export_data_z.append(r_set_2)
        export_data_x.append(Q_dot)
        export_data_y.append(Te)
    elif selection == 1 : 
        T_diff = init_T_i - T_i
        export_data_x.append(x_data)
        export_data_y.append(T_diff)
    elif selection == 2 :
        export_data_x.append(x_data)
        export_data_y.append(velo_fluid_new)
    elif selection == 3 :
        export_data_x.append(x_data_2)
        export_data_y.append(Te)
    elif selection == 4 :
        export_data_x.append(x_data)
        export_data_y.append(Q_dot)
        export_data_z.append([m_dot,fluid_cp])
    elif selection == 5 :
        T_diff = init_T_i - T_i
        export_data_x.append(x_data_2)
        export_data_y.append(T_diff)
    elif selection == 100 :
        export_data_x = Te
    return export_data_x, export_data_y, export_data_z

def get_random_color():
    r = random.random()
    g = random.random()
    b = random.random()
    return (r, g, b)

def graph_data (length, section_number ,layer_diameter_set, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) : 
    x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_diameter_set, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)

    if selection == 0 : 
        plot_x = [0]
        plot_x.extend(layer_diameter_set)
        plot_x.append(layer_diameter_set[-1]*1.25)
        y_axis = core_HT.radial_analysis(z_data[0], y_data[0], x_data[0]) # inside determined the placement to measure
        x_axis = plot_x
        xlabel = "Radial (m)"
        ylabel = "Temeprature (°C)"
        title = "Radial Temperature Profile"
    elif selection == 1 :
        x_axis = x_data
        y_axis = y_data
        xlabel = "PipeLength (m)"
        ylabel = "Temeprature Difference (°C)"
        title = "Temperature Difference across Pipe Length"
    elif selection == 2 : 
        x_axis = x_data
        y_axis = y_data
        xlabel = "PipeLength (m)"
        ylabel = "Flow Velocity (ms-1)"
        title = "Operation Velocity across Pipe Length"
    elif selection == 3 :
        x_axis = x_data
        y_axis = y_data
        xlabel = "PipeLength (m)"
        ylabel = "Temeprature (°C)"
        title = "Axial Temperature Profile"
    elif selection == 4: 
        x_axis = x_data
        y_axis = y_data
        xlabel = "PipeLength (m)"
        ylabel = "Heat Transfer (W)"
        title = "Heat Trasfer across Pipe Length"
    elif selection == 5 :
        x_axis = x_data
        y_axis = y_data
        xlabel = "PipeLength (m)"
        ylabel = "Temeprature Difference (°C)"
        title = "Temperature Difference across Pipe Length"

    return x_axis, y_axis, xlabel, ylabel, title

def case_01 (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) :
    x_plot, y_plot, label_x, label_y, title = graph_data(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)

    if selection == 0: 
        pmat = []
        layer = layer_thickness
        layer.append(layer_thickness[-1]*1.25)
        for i in range (len(insulator_selection)):
            pmat.append(material.data_set_insulator_mat[insulator_selection[i]])
        lgd = ["Pipe Inner Wall", "Pipe Outer Wall"]
        for i in range (len(layer_thickness)-3): 
            lgd.append("Insulator layer " + str(i+1) + " - " + str(pmat[i]))
        lgd.append ("Ambient")
        for i in range (len(layer)) :
            plt.axvline (layer[i],color=get_random_color(),linestyle='--',label = lgd[i],alpha = 0.5)
        labels = "Temperature Profile"
        plt.plot(x_plot,y_plot,marker="o",color='orange',zorder = 2)
        for i, j in zip(x_plot,y_plot) :
            plt.annotate(str(round(j,2))+" °C", xy=(i,j), xytext=(5,5), textcoords='offset points')
    else :
        labels = "Model"
    plt.plot(x_plot,y_plot,label=labels)
    plt.xlabel(label_x)
    plt.ylabel(label_y)
    plt.title(title)
    plt.legend()
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show()

def __sign(value):
    if value == 0:
        return 0
    elif value > 0:
        return 1
    elif value < 0:
        return 0

def __find_apparent_order_iteratively(e21, e32, r21, r32):
    eps = 1
    iteration = 1
    max_iteration = 1000
    norm = 1
    p = 2
    while eps > 1e-6:
        p_old = p
        s = __sign(e32 / e21)
        q = np.log((pow(r21, p) - s) / (pow(r32, p) - s))
        p = (1.0 / np.log(r21)) * abs(np.log(abs(e32 / e21)) + q)

        residual = p - p_old
        if iteration == 1:
            norm = residual
        iteration += 1
        eps = residual / norm
        if iteration == max_iteration:
            print('WARNING: max number of iterations reached for calculating apparent order p ...')
            break
    return p

def grid_convergence_index(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) : 
    r_32 = 2
    r_21 = 2
    Safety_factor = 1.25
    desired_GCI = 0.05 
    
    section_set = [section_number]
    h_3 = length/section_number
    h_2= h_3 / r_32 
    section_set.append(int(length/h_2))

    h_1 = h_2 / r_21
    section_set.append(int(length/h_1))

    te_set = []
    for i in range (len(section_set)) :
        te, x, y = core_HT.sectional_analysis(length, section_set[i] ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
        te_set.append(float(te))
    print(te_set)
    print(section_set)
    if r_32 == r_21 :
        p_value = np.log(abs(te_set[0]-te_set[1])/abs(te_set[-2]-te_set[-1])) / np.log(r_21)
    else : 
        p_value = __find_apparent_order_iteratively(abs(te_set[-2]-te_set[-1]), abs(te_set[0]-te_set[1]),r_21,r_32)
    print("P value = " + str(p_value))

    CR = (te_set[1] - te_set[-1]) / (te_set[0]-te_set[1])
    if 0 <= CR and CR < 1 :
        print("Monotonic Convergence")
    else :
        print("Failed")
    print("CR-Value : " + str(CR))

    #extrapolate_1 = np.power(r_21,p_value) * (te_set[-1] - te_set[1]) / (np.power(r_21,p_value)-1)
    extrapolate_2 = te_set[-1] + (te_set[-1] - te_set[1]) / (np.power(r_21,p_value)-1)
    #print("Extrapolata Value_Web = " + str(extrapolate_1))
    print("Extrapolata Value_GPT = " + str(extrapolate_2))

    GCI_32 = (Safety_factor * abs((te_set[1] - te_set[0])/te_set[1])) / (np.power(r_32,p_value)-1)
    GCI_21 = (Safety_factor * abs((te_set[-1] - te_set[1])/te_set[-1])) / (np.power(r_21,p_value)-1)
    print("GCI_32 = " + str(GCI_32))
    print("GCI_21 = " + str(GCI_21))

    AR = GCI_32 / (np.power(r_21,p_value) * GCI_21)
    print("AR value = " + str(AR))

    #expected_r = np.power((desired_GCI/GCI_21),1/p_value)
    #print("expected r-value = " + str(expected_r))

def plot_convergence (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) :
    section_number_set = [4000, 2000, 1600, 1000, 800, 400, 200, 100, 90, 60, 40, 20, 10, 4, 2, 1]
    selection = 3
    x_data_set = []
    y_data_set = []
    labels = []
    for i in range (len(section_number_set)) :
        x_data, y_data, z_data = core_HT.sectional_analysis (length, section_number_set[i] ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
        x_data_set.append(x_data)
        y_data_set.append(y_data)
        labels.append("N_Cells = "+str(section_number_set[i]))
    for i in range (len(x_data_set)): 
        plt.plot(x_data_set[i],y_data_set[i],label = labels[i], color=get_random_color())
    
    plt.xlabel ("PipeLength (m)")
    plt.ylabel ("Temeprature (°C)")
    plt.title ("Axial Temperature Profile (Convergence Index Study)")
    plt.legend()
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show ()

def model_compare (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) :
    selection = 3
    condition = [True, False]
    plot_x = []
    plot_y = []
    label = []
    for j in range (len(condition)) :
        model[1] = condition[j]
        for i in range (len(condition)) :
            model[0] = condition[i]
            x_axis , Te, non = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
            plot_x.append(x_axis)
            plot_y.append(Te)
            label.append("Model Sectional Analysis = " + str(model[0]) + ", Model Natural Convection = " + str(model[1]))

    for i in range (len(plot_x)): 
        plt.plot(plot_x[i], plot_y[i],label = label[i])
    plt.plot([],[], label = "h_amb = 20 for Off Model Natural Convection" ,color = "none")
    
    plt.xlabel("PipeLength (m)")
    plt.ylabel("Temeprature (°C)")
    plt.title("Axial Temperature Profile")
    plt.legend()
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show()

def wang_valid(valid_selection) : 
    T_i = 100
    velo_fluid = 2.5
    T_amb = 0
    insulator_selection = [20]
    fluid_selection = 1
    pipe_selection = 2
    length = 50000
    section_number = 100
    model = [True, False]
    selection = 5

    inner_pipe = 1
    pipe_thick = 0.005
    insu_thick = 0.05
    layer_thickness = [inner_pipe, inner_pipe+pipe_thick*2, (inner_pipe+pipe_thick*2)+insu_thick]

    var = [layer_thickness,T_i,velo_fluid,T_amb]
    var_selection = 0

    if valid_selection == "3a" : 
        x_plot,y_plot,x2_plot,y2_plot,label_x,label_y, legend_1,legend_2,title,external_data = Validation.valid_wang_fig_3a ()
        var_selection = 0
        inner_pipe = [0.4, 1, 2]
        group = []
        for i in range (len(inner_pipe)): 
            layer_thickness = [inner_pipe[i], inner_pipe[i]+pipe_thick*2, (inner_pipe[i]+pipe_thick*2)+insu_thick]
            group.append(layer_thickness)
        var_set = group
        legend_model = ["DN=0.4m (Model)", "DN=1.0m (Model)","DN=2.0m (Model)"]
        legend_model_2 = []
        for i in range (len(legend_model)) :
            legend_model_2.append(str(legend_model[i]) + " With Natural Convection")
        for i in range (len(legend_model)) :
            legend_model[i] = (str(legend_model[i]) + " h_amb = 20")
        
    elif valid_selection == "3b" :
        x_plot,y_plot,x2_plot,y2_plot,label_x,label_y, legend_1,legend_2,title,external_data = Validation.valid_wang_fig_3b ()
        var_selection = 1
        var_set = [70, 100, 150]
        legend_model = ["Ti=70°C (Model)", "Ti=100°C (Model)","Ti=150°C (Model)"]
        legend_model_2 = []
        for i in range (len(legend_model)) :
            legend_model_2.append(str(legend_model[i]) + " With Natural Convection")
        for i in range (len(legend_model)) :
            legend_model[i] = (str(legend_model[i]) + " h_amb = 20")

    elif valid_selection == "3c" :
        x_plot,y_plot,x2_plot,y2_plot,label_x,label_y, legend_1,legend_2,title,external_data = Validation.valid_wang_fig_3c ()
        var_selection = 2
        var_set = [1.0, 2.5, 5.0]
        legend_model = ["v=1.0ms-1 (Model)", "v=2.5ms-1 (Model)","v=5.0ms-1 (Model)"]
        legend_model_2 = []
        for i in range (len(legend_model)) :
            legend_model_2.append(str(legend_model[i]) + " With Natural Convection")
        for i in range (len(legend_model)) :
            legend_model[i] = (str(legend_model[i]) + " h_amb = 20")

    elif valid_selection == "3d" :
        x_plot,y_plot,x2_plot,y2_plot,label_x,label_y, legend_1,legend_2,title,external_data = Validation.valid_wang_fig_3d ()
        var_selection = 3
        var_set = [-40, 0, 40]
        legend_model = ["Ts=-40°C (Model)", "Ts=0°C (Model)","Ts=40°C (Verified Model)"]
        legend_model_2 = []
        for i in range (len(legend_model)) :
            legend_model_2.append(str(legend_model[i]) + " With Natural Convection")
        for i in range (len(legend_model)) :
            legend_model[i] = (str(legend_model[i]) + " h_amb = 20")
    
    x_graph = []
    y_graph = []
    for i in range (3) : 
        var[var_selection] = var_set[i]
        x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,var[0], fluid_selection, pipe_selection, insulator_selection,var[1],var[3],var[2], selection, model)
        x_graph.append(x_data)
        y_graph.append(y_data)
    
    model = [True, True]
    x_graph_2 = []
    y_graph_2 = []
    for i in range (3) : 
        var[var_selection] = var_set[i]
        x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,var[0], fluid_selection, pipe_selection, insulator_selection,var[1],var[3],var[2], selection, model)
        x_graph_2.append(x_data)
        y_graph_2.append(y_data)

    color = ['#b0354c','#07a466','#7548ab']
    for i in range (3) :   
        plt.plot(x_graph[i],y_graph[i],label = legend_model[i],color=color[i])
        plt.plot(x_graph_2[i],y_graph_2[i],'--',label = legend_model_2[i],color=color[i])
        plt.plot(x_plot[i],y_plot[i], ':', label = legend_1 [i],color=color[i])
        a = y_graph[i]
        b = y_graph_2[i]
        c = y_plot[i]
        
        print(a[-1])
        print(b[-1])
        print(c[-1])
    
    plt.xlabel(label_x)
    plt.ylabel(label_y)
    plt.title(title)
    plt.legend()
    plt.grid(True)
    plt.tight_layout()
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show()

def material_naming (selection_pipe, selection_fluid) : 
    if selection_pipe == 1 :
        material = "Copper"
    elif selection_pipe == 2 :
        material = "Aluminum"
    elif selection_pipe == 3: 
        material = 'Iron'
    
    if selection_fluid == 1 :
        fluid = "Sat_water"
    elif selection_fluid == 2 :
        fluid = "Engine Oil"
    elif selection_fluid == 3 :
        fluid = "Methane Gas"
    elif selection_fluid == 4 : 
        fluid = "Crude Oil"
    
    return material, fluid 

def parametric_study(study_num) :
    length = 50000
    section_number = 100

    fluid_selection = 4
    pipe_selection = 3
    insulator_selection = [12]

    pipe_outer = 0.1683
    pipe_inner = pipe_outer - 0.01097*2
    insulator_1 = pipe_outer + 0.1
    layer_thickness = [pipe_inner, pipe_outer, insulator_1]

    T_i = 250
    T_amb = 27
    velo_fluid = 2.5

    model = [True, True]

    selection = 3

    if study_num == 1 : 
        fluid_selection = [1, 2, 4]
        pipe_selection = [1, 2, 3]

        x_plot = []
        y_plot = []
        lgd= []
        for i in range (len(fluid_selection)) :
            for j in range (len(pipe_selection)) :
                x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection[i], pipe_selection[j], insulator_selection,T_i,T_amb,velo_fluid, selection, model)
                x_plot.append(x_data)
                y_plot.append(y_data)
                pipe, fluid = material_naming(pipe_selection[j],fluid_selection[i])
                lgd.append(str(fluid) + ", " + str(pipe))
        for i in range (len(x_plot)): 
            plt.plot (x_plot[i],y_plot[i], label=lgd[i])
        
        ylabel = "Temeprature(°C)"
    
    elif study_num == 2 :
        x_plot = []
        y_plot = []
        lgd= []
        T_i = []

        selection = 1
        for i in range (10) :
            T_i.append(60+i*20)
        
        for i in range (len(T_i)) :
            x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i[i],T_amb,velo_fluid, selection, model)
            x_plot.append(x_data)
            y_plot.append(y_data)
            lgd.append("Ti = " + str(T_i[i]) + "°C")
            print(y_data[-1])
        
        for i in range (len(x_plot)): 
            plt.plot (x_plot[i],y_plot[i], label=lgd[i])
        
        if selection == 3 : 
            ylabel = "Temeprature Exit (°C)"
        elif selection == 1 : 
            ylabel = "Temeprature Difference(°C)"

    elif study_num == 3 :
        x_plot = []
        y_plot = []
        lgd= []
        T_amb = []

        selection = 1
        for i in range (10) :
            T_amb.append(20+i*2)
        
        for i in range (len(T_amb)) :
            x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb[i],velo_fluid, selection, model)
            x_plot.append(x_data)
            y_plot.append(y_data)
            lgd.append("Tamb = " + str(T_amb[i]) + "°C")
            print(y_data[-1])
        
        for i in range (len(x_plot)): 
            plt.plot (x_plot[i],y_plot[i], label=lgd[i])
        
        if selection == 3 : 
            ylabel = "Temeprature Exit (°C)"
        elif selection == 1 : 
            ylabel = "Temeprature Difference(°C)"

    elif study_num == 4 : 
        x_plot = []
        y_plot = []
        lgd= []
        v = []

        selection = 1
        for i in range (10) :
            v.append(0.5+i*0.5)
        
        for i in range (len(v)) :
            x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,v[i], selection, model)
            x_plot.append(x_data)
            y_plot.append(y_data)
            lgd.append("Initial Fluid Velocity = " + str(v[i]) + "ms-1")
            print(y_data[-1])
        
        for i in range (len(x_plot)): 
            plt.plot (x_plot[i],y_plot[i], label=lgd[i])
        
        if selection == 3 : 
            ylabel = "Temeprature Exit (°C)"
        elif selection == 1 : 
            ylabel = "Temeprature Difference(°C)"
    
    elif study_num == 5 :
        x_plot = []
        y_plot = []
        lgd= []
        selection = 1

        insulator_thickness = [0.1, 0.2, 0.4]
        insulator_layer = [1, 2, 3]

        for i in range (len(insulator_thickness)) :
            layer_thickness_1 = [pipe_inner, pipe_outer, insulator_1]
            insulator_selection = [12]
            for j in range (len(insulator_layer)) : 
                layer_thickness_1 [2] = layer_thickness_1[1] + insulator_thickness[i]*2
                x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness_1, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
                layer_thickness_1.append(layer_thickness_1[-1] + insulator_thickness[i]*2)
                insulator_selection.append(12)
                print(y_data[-1])
                x_plot.append(x_data)
                y_plot.append(y_data)
                lgd.append("Insulation Thickness = " + str(insulator_thickness[i]) + "m, Insulation Layer = " + str(insulator_layer[j]))
        if selection == 3 : 
            ylabel = "Temeprature Exit (°C)"
        elif selection == 1 : 
            ylabel = "Temeprature Difference(°C)"
        
        for i in range (len(x_plot)) :
            plt.plot (x_plot[i], y_plot[i],label= lgd[i])

    elif study_num == 6 :
        x_plot = []
        y_plot = []
        lgd= []
        selection = 2

        pipe_outer = [0.1413, 0.1683, 0.2191]
        pipe_thickness = [0.01097, 0.01270, 0.01826]

        pipe_outer_word = ["DN125", "DN150", "DN200"]
        pipe_thickness_word = [0.01097, 0.01270, 0.01826]

        for i in range (len(pipe_outer)) :
            layer_thickness = [pipe_inner, pipe_outer, insulator_1]
            for j in range (len(pipe_thickness)) :
                layer_thickness[1] = pipe_outer[i]
                layer_thickness[0] = pipe_outer[i] - pipe_thickness[j]
                x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
                x_plot.append(x_data)
                y_plot.append(y_data)
                lgd.append(str(pipe_outer_word[i])+ ", P.Thick = " + str(pipe_thickness_word[j]))
                print(y_data[-1])
        
        
        if selection == 3 : 
            ylabel = "Temeprature Exit (°C)"
        elif selection == 1 : 
            ylabel = "Temeprature Difference(°C)"
        elif selection == 2:
            ylabel = "Flow Velocity (ms-1)"

        for i in range (len(x_plot)) :
            plt.plot (x_plot[i], y_plot[i],label= lgd[i])



    xlabel = "PipeLength (m)"
    title = "Axial Temperature Profile"
    title = "Operation Velocity across Pipe Length"
    
    plt.xlabel(xlabel)
    plt.ylabel(ylabel)
    plt.title(title)
    plt.legend()
    plt.grid(True)
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show()

def basic_mix_mode (selection_1, selection_2, length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model ) :
    data_1 = []
    label_1 = []
    selection = selection_1
    x_plot, y_plot, label_x, label_y, title = graph_data(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    data_1 = y_plot
    label_1 = label_y

    data_2 = []
    label_2 = []
    selection = selection_2
    x_plot, y_plot, label_x, label_y, title = graph_data(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    data_2 = y_plot
    label_2 = label_y

    plt.plot(data_1,data_2)
    plt.title(str(label_2) + " Vs " + str(label_1))
    plt.xlabel(label_1)
    plt.ylabel(label_2)
    plt.grid(True)
    plt.show()

def external_pipe_radial (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model) : 
    selection = 0
    delta_pipe = length / section_number
    x_data, y_data, z_data = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    colors = ['r', 'b', 'g']
    stage = [0, len(x_data)/2, len(x_data)-1]
    for m in range (3) :
        num = int(stage [m] )
        plot_x = [0]
        plot_x.extend(layer_thickness)
        plot_x.append(layer_thickness[-1]*1.25)
        y_axis = core_HT.radial_analysis(z_data[num], y_data[num], x_data[num]) # inside determined the placement to measure
        x_axis = plot_x
        xlabel = "Radial (m)"
        ylabel = "Temeprature (°C)"
        title = "Radial Temperature Profile"

        x_plots, y_plots, label_x, label_y, title = x_axis, y_axis, xlabel, ylabel, title 
        pmat = []
        layer = []
        for i in range (len(x_plots)-2) :
            layer.append(x_plots[i+1]/2)
        layer.append(x_plots[-1]/2)
        for i in range (len(insulator_selection)):
            pmat.append(material.data_set_insulator_mat[insulator_selection[i]])
        lgd = ["Pipe Inner Wall", "Pipe Outer Wall"]
        for i in range (len(x_plots)-4): 
            lgd.append("Insulator layer " + str(i+1) + " - " + str(pmat[i]))
        lgd.append ("Ambient")
        

        x_plot = []
        y_plot = []
        for i in range (len(layer_thickness)-1):
            x_plot_1, y_plot_1 = core_HT.plot_radial(layer_thickness[i+1],layer_thickness[i],y_plots[i+1],y_plots[i+2])
            x_plot.append(x_plot_1)
            y_plot.append(y_plot_1)

        for i in range (len(x_plot)) :
            plt.plot(x_plot[i],y_plot[i],color= colors[m])

        scnd_x = [x_plots[0]/2, x_plots[1]/2]
        scnd_y = [y_plots[0], y_plots[1]]

        scnds_x = [x_plots[-1]/2, x_plots[-2]/2]
        scnds_y = [y_plots[-1], y_plots[-2]]
        plt.plot(scnd_x,scnd_y, color= colors[m],label="Length = " + str(num*delta_pipe))
        plt.plot(scnds_x,scnds_y, color= colors[m])

        for i in range (len(x_plots)) :
            x_plots[i] = x_plots[i]/2
        plt.plot(x_plots,y_plots,"o",color = "orange")
        for i, j in zip(x_plots,y_plots) :
            plt.annotate(str(round(j,2))+" °C", xy=(i,j), xytext=(5,5), textcoords='offset points')

    for i in range (len(layer)) :
            plt.axvline (layer[i],color=get_random_color(),linestyle='--',label = lgd[i],alpha = 0.5)

    plt.xlabel(label_x)
    plt.ylabel(label_y)
    plt.title(title)
    plt.legend()
    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    plt.show()

def anova_data_prep () :
    filename = "C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Data\\library_sheet.xlsx"
    sheet_name = "Datas"
    df = pd.read_excel(filename, sheet_name)
    section_number = 100
    model = [True, True]
    insulator_selection = [12]
    selection = 1

    Ti_set = np.array(df['Column1'])
    Tamb_set = np.array(df['Column2'])
    velo_set = np.array(df['Column3'])
    fluid_T_set = np.array(df['Column4'])
    pipe_T_set = np.array(df['Column5'])
    insu_thick_set = np.array(df['Column6'])
    pipe_dia_set = np.array(df['Column7'])
    pipe_thick_set = np.array(df['Column8'])
    length_set = np.array(df['Column9'])

    ans = []
    num = 5
    print("predicted system time = " + str(num*1.5/60) + " min")
    for i in range (num) :
        layer_thickness = [pipe_dia_set[i]-pipe_thick_set[i],pipe_dia_set[i],pipe_dia_set[i]+insu_thick_set[i]]
        fluid_selection = fluid_T_set[i]
        pipe_selection = pipe_T_set[i]
        T_i = Ti_set[i]
        T_amb = Tamb_set[i]
        velo_fluid = velo_set[i]
        length = length_set[i]
        data_x, data_y, data_z = core_HT.sectional_analysis(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
        ans.append(data_y[-1])
    
    
    wb = Workbook()
    wb = load_workbook("C:\\Users\\zyneo\\OneDrive\\Desktop\\Results\\temp.xlsx")
    border_thin = Side(style='thin')
    ws = wb["Sheet1"]

    for i in range (len(ans)): 
        ws.cell(row = i+2, column = 2).value = ans[i]
    
    wb.save("C:\\Users\\zyneo\\OneDrive\\Desktop\\Results\\temp.xlsx") 

    playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')


if __name__ == "__main__" : 
    print("Running")
    
    length = 50000
    section_number = 100

    inner_pipe =0.14636
    insulator = 0.2183
    insulator2 = 0.2683
    layer_thickness = [inner_pipe, 0.1683, insulator]
    fluid_selection = 4 # 1 = satwater, 2 = engine oil, 3 = methane gas, 4 = Crude Oil
    pipe_selection = 3 # 1 = copper , 2 = aluminum, 3 = iron
    insulator_selection = [12]

    T_i = 250
    T_amb = 27

    velo_fluid = 2.5

    selection = 62
    model = [True, True]

    if selection == 100 :
        grid_convergence_index(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
        playsound('C:\\Users\\zyneo\\OneDrive\\Desktop\\P5_Neoh\\1 Main File\\Code\\done bgm.mp3')
    elif selection == 101 : 
        plot_convergence (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    elif selection == 50 :
        model_compare(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    elif selection == 51 :
        wang_valid("3c")
    elif selection == 60 :
        parametric_study(6)
    elif selection == 61 :
        basic_mix_mode (2,4, length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model )
    elif selection == 62 :
        external_pipe_radial(length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)
    elif selection == 63 : 
        anova_data_prep()
    elif selection == 64 :
        a = 1
    else :
        case_01 (length, section_number ,layer_thickness, fluid_selection, pipe_selection, insulator_selection,T_i,T_amb,velo_fluid, selection, model)

    print("Complete")